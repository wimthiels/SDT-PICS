#!/usr/bin/python3.7
'''
Created on 17 Aug 2019
@author: wimth
run in bash
modifies a XML based on an excel.  This excel contains the inputxml location, outputlocation, and all the different fields to be modified.
The original xml is left untouched
every sheet in the excel corresponds to a runnumber
'''
'''
'''
import xml.etree.ElementTree as ET
from pathlib import Path
import sys,os,shutil,re
import openpyxl as xl
import pandas as pd

#sys.arg : $INPUT_CSV $XML_FILE $runID

def read_excel():
	INPUT_CSV = sys.argv[1] if len(sys.argv)>1 else 'testinfra.xlsx' 
	wb = xl.load_workbook(filename = INPUT_CSV)
	
	# READ CONFIG TAB
	ws_config = wb["config"]
	d_config = {}  
	
	for xrow in ws_config.iter_rows(min_row=1, min_col=1, max_row=ws_config.max_row, max_col=ws_config.max_column):
		if xrow[0].value.endswith('_xml') and xrow[1].value:
			d_config[xrow[0].value] = Path(xrow[1].value)
		else:
			d_config[xrow[0].value] = xrow[1].value
	
	if len(sys.argv)>3:
		d_config['NB_RUN'] = str(sys.argv[3])  #sys.argv takes precedence over excel

	# READ DATA TAB
		# first load testcase variables (= static variables)
	try:
		df_tc_info = pd.read_excel(INPUT_CSV,sheet_name='tc_info',converters={'suffix':str,'offset':str,'replicate':str,'time':str,'nb_stack_raw':str})
		sel = df_tc_info['testcase']==d_config['NB_RUN']
		df_tc_info = df_tc_info[sel].reset_index()
	except:
		df_tc_info = pd.DataFrame({})

		# load tab for this testcase
	try:
		ws_data = wb["{0}".format(d_config['NB_RUN'])]
	except Exception:
		ws_data=None

	if not ws_data:
		try:
			d_config['NB_RUN'] = df_tc_info['use_tab'][0]
			ws_data = wb[d_config['NB_RUN']]
		except Exception as e:
			print(f'Loading parameter data from excel failed : exitting {e}')
			exit(-1)
		

		# extract dynamic variables (=variables derived from fields on the testcase tab) and 
	output_folder=None
	init_folder=None
	sub_folder=""
	switch_tab=[]
	for xrow in ws_data.iter_rows(min_row=1, min_col=1, max_row=ws_data.max_row, max_col=ws_data.max_column):

		if xrow[0].value.endswith('OUTPUT_FOLDER_ROOT'):
			output_folder = Path(fill_in_variables(xrow[1].value,df_tc_info))

		if xrow[0].value.endswith('OUTPUT_FOLDER_ROOT_SUB'):
			sub_folder = fill_in_variables(xrow[1].value,df_tc_info)

		if xrow[0].value.endswith('<INIT_FOLDER>'):
			init_folder = Path(fill_in_variables(xrow[1].value,df_tc_info))

		if xrow[0].value =='<SWITCH_TAB>':
			if xrow[1].value:
				#switch_tab = xrow[1].value.split(";")
				switch_tab = fill_in_variables(xrow[1].value,df_tc_info).split(";")


		
	switch_tab.append(d_config['NB_RUN'])

	return wb,d_config,output_folder,init_folder,sub_folder,switch_tab,INPUT_CSV,df_tc_info
	
def fill_in_variables(xrow_value,df_tc_info):
	"""
	$ variables are defined in other fields of the tab (=dynamic parameters), used for path resolving
	# variables are listed in the tc_info tab and can also be interpreted as expressions 

	"""
	s_new_value = str(xrow_value)   #turn everything to string

	for substitution in re.findall(r"#(.+?)#",str(xrow_value)):
		s_new_value=s_new_value.replace(f"#{substitution}#",str(df_tc_info[substitution][0]))
		if substitution == 'offset':
			s_new_value = str(eval(s_new_value))

	if "$DATA_FOLDER" in s_new_value:
		s_new_value = s_new_value.replace("$DATA_FOLDER", str(data_folder))
		
	if "$SCRIPT_FOLDER" in s_new_value:
		s_new_value = s_new_value.replace("$SCRIPT_FOLDER", str(script_folder))



	return s_new_value

def modify_based_on_excel_tab(ws_data,d_config,df_tc_info):
	#modify based on excel
	global flag_cancel_init_folder
	for xrow in ws_data.iter_rows(min_row=1, min_col=1, max_row=ws_data.max_row, max_col=ws_data.max_column):   
		try:
			if xrow[0].value.strip() in ['<INIT_FOLDER>','<SWITCH_TAB>']:
				continue
				
			xml_element = root.findall(xrow[0].value.strip())[0]  #gives back a list in theory, we only use first match
			
			s_new_value = fill_in_variables(xrow[1].value,df_tc_info)

			if s_new_value == 'blanc':
				xml_element.set('value','')
			else:
				xml_element.set('value',s_new_value)  #this will set @value (if there is an text entry it will create a orderedDict eg OrderedDict([('@value','TL11_emb1of1_t1.tif'),('#text', 'file1')])),
				
			if ("ix_restart" in xrow[0].value):
				if s_new_value not in ["","0",0] and d_config.get('init_folder'):
					flag_cancel_init_folder = True
				else:
					flag_cancel_init_folder = False


			if verbose:print(xrow[0].value.strip(),"=>",xml_element.get('value'))
		except:
			if verbose:print("{0} : xml value not found".format(xrow[0]))
		#to be set

	return 

	
	
#MAIN----------------------------------------------------------------------------------------

#$INPUT_CSV $XML_OUTPUT $runID $XML_INPUT $DATA_FOLDER $SCRIPT_FOLDER
input_xml     = Path(sys.argv[4]) if len(sys.argv)>4 else Path("/mnt/c/Users/wimth/OneDrive/SYNC/PARAMS.xml")
output_xml_folder    = Path(sys.argv[2]) if len(sys.argv)>2 else Path("/mnt/c/Users/wimth/OneDrive/SYNC/XML")
data_folder   = Path(sys.argv[5]) if len(sys.argv)>5 else Path("/mnt/f/Downloads/ROET/testinfra")
script_folder = Path(sys.argv[6]) if len(sys.argv)>6 else Path("/mnt/c/Users/wimth/OneDrive/SYNC")
param_folder = Path(sys.argv[7]) if len(sys.argv)>7 else script_folder
restart_module = int(sys.argv[8]) if len(sys.argv)>8 else 0
 
output_xml = output_xml_folder / "testinfra_{0}.xml".format(str(sys.argv[3]))


verbose=True
wb,d_config,output_folder,init_folder,sub_folder,switch_tab,INPUT_CSV,df_tc_info = read_excel()


#compose XML from base state and excel-modifications
if d_config['input_xml']: # alternative XML
	print("Using alternate XML as input source : str(d_config['input_xml'])=",str(d_config['input_xml']))  #if an alternate xml file is filled in this, will be picked (no modifications done)
if d_config['input_xml']:
	tree = ET.parse(str(d_config['input_xml']))
	print("***RESTARTING FROM ",str(d_config['input_xml']),"***")
else:  # standard PARAM FILE IS MODIFIED based on excel
	tree = ET.parse(input_xml)
	root = tree.getroot()
	
	flag_cancel_init_folder = False
	for tab_i in switch_tab: 
		if verbose:print(">>modifying PARAM state with tab={0} from {1}<<".format(tab_i,INPUT_CSV))
		ws_data = wb["{0}".format(tab_i)]
		modify_based_on_excel_tab(ws_data,d_config,df_tc_info)
	if flag_cancel_init_folder:
		d_config.pop('init_folder',None)
		print('Info : directory not initialized because of checkpoint restart')

	#store general data
	xml_element = root.findall('./MAIN/paths/OUTPUT_FOLDER_ROOT')[0]
	# xml_element.set('value',str(output_xml.parent))
	xml_element.set('value',str(output_folder))
	xml_element = root.findall('./RAM/script_folder')[0]  #Note: if xml-modifier is run under docker, then the default script folder will be docker (only use docker_ic to force docker script folder)
	xml_element.set('value',str(script_folder))
	xml_element = root.findall('./RAM/data_folder')[0]
	xml_element.set('value',str(data_folder))



# #actions
  #init folder
if d_config.get('init_folder'):
	if not init_folder:
		if output_folder:
			if sub_folder:
				init_folder = output_folder/sub_folder
			else:
				init_folder = output_folder

	if init_folder:
		if restart_module > 1:
			print("not deleting init-folder because resuming from module {0}".format(restart_module))
		else:
			if init_folder.exists():
				if 'OUTPUT' not in init_folder.parts or init_folder.parts[-1] == 'OUTPUT':
					print('INIT_FOLDER does not have OUTPUT in the pathname (or OUTPUT is the root folder).  Are you really sure you want to remove this folder ? if so, do it manually')

				else:
					shutil.rmtree(str(init_folder),ignore_errors=True)  
					if verbose:print(str(init_folder), " is removed by xml-modifier")
else:
	print('Folders are not intialized because init_folder on the config-tab is set to off - ', d_config.get('init_folder'))


  #create folders
if not output_folder:
	output_folder = init_folder
	if not output_folder:
		print('WARNING : output_folder nor INIT_FOLDER is defined.  add to excel and restart')
if sub_folder:
	(output_folder/sub_folder/'_LOG'/'modules_snapshot').mkdir(parents=True,exist_ok=True)
	print((output_folder/sub_folder/'_LOG'/'modules_snapshot')," has been created")
else:
	(output_folder/'_LOG'/'modules_snapshot').mkdir(parents=True,exist_ok=True)
	print(str(output_folder/'_LOG'/'modules_snapshot')," has been created")
			
  #store param xml file
if output_xml.exists():os.remove(output_xml)
output_xml_folder.mkdir(parents=True,exist_ok=True)
tree.write(str(output_xml))
print('output_xml is written :',output_xml)

#pre resolve variables ad-hoc (the param file needs some variables to be fully resolved to allow for extraction by a simple grep or text search outside a python context)
from param_XML import Param_xml
param_xml = Param_xml.get_param_xml(['',output_xml], l_main_keys=['body'],verbose=True)
param_xml.add_prms_to_dict(l_keys=['sphere_meshing']) 
for key,item in param_xml.prm.items():
	item = str(item)
	if key in ['shrink_factor','threshold_radius_micron']:
		search_key = './sphere_meshing/parms/' +  key
	else:
		search_key = './sphere_meshing/paths/' +  key
	xml_element = root.findall(search_key)[0]
	if xml_element:
		xml_element.set('value',item)

tree.write(str(output_xml))  #update the xml-file




